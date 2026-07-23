"""
io_load.py — Resilient ingestion of the THREE raw datasets (Step 1).

Project: Predicting Bagrut Success from Municipal Socioeconomics and
         School-Level Institutional Resources
Authors: Yousef Shihade & Shada Esawi

Responsibilities
----------------
* Dataset 1 — Bagrut grades CSV: handle the UTF-8 *BOM* so the first header is
  not corrupted ("grade" -> "﻿grade").
* Dataset 2 — CBS socioeconomic index XLSX: the real header sits on row index 10;
  Hebrew headers are renamed to canonical English names and the ``..`` "unranked"
  placeholders are coerced to NaN.
* Dataset 3 — Ministry of Education budget XLSX : the workbook
  ships a malformed ``styles.xml`` (non-aRGB theme colours) which makes a vanilla
  ``openpyxl.load_workbook`` raise ``ValueError: Colors must be aRGB hex values``.
  We monkeypatch openpyxl's ``RGB`` descriptor with a lenient validator *before*
  opening the file, then drop the grand-totals row and whitespace-normalise the
  Hebrew headers.

These functions return raw (un-cleaned) DataFrames; text standardisation lives in
``clean_text.py`` so ingestion and cleaning stay independently testable.
"""
from __future__ import annotations

import re
import warnings
from pathlib import Path
from typing import Any

import pandas as pd
import yaml

ROOT = Path(__file__).resolve().parents[1]
CONFIG_PATH = ROOT / "config.yaml"

_WS_RE = re.compile(r"\s+")


def load_config(path: Path | str = CONFIG_PATH) -> dict[str, Any]:
    """Load ``config.yaml`` as a plain dict."""
    with open(path, "r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def resolve(rel_path: str) -> Path:
    """Resolve a config-relative path against this step folder."""
    return (ROOT / rel_path).resolve()


def _norm_header(value: Any) -> Any:
    """Collapse internal whitespace + trim (Hebrew headers carry double spaces)."""
    if value is None:
        return value
    return _WS_RE.sub(" ", str(value)).strip()


# --------------------------------------------------------------------------- #
# Dataset 1 — Bagrut grades                                                    #
# --------------------------------------------------------------------------- #
def load_bagrut(cfg: dict[str, Any] | None = None) -> pd.DataFrame:
    """Load ``israel_bagrut_averages.csv`` (UTF-8 *with BOM*)."""
    cfg = cfg or load_config()
    df = pd.read_csv(resolve(cfg["paths"]["raw_bagrut"]),
                     encoding=cfg["bagrut"]["encoding"])
    # Belt-and-braces: drop any residual BOM/whitespace from the labels.
    df.columns = [str(c).replace("﻿", "").strip() for c in df.columns]
    return df


# --------------------------------------------------------------------------- #
# Dataset 2 — CBS socioeconomic index                                          #
# --------------------------------------------------------------------------- #
def _resolve_rename_prefix(raw_columns: list[str], rename_cfg: dict[str, str]) -> dict[str, str]:
    """{raw_header -> canonical} using PREFIX matching.

    The CBS cluster header carries a trailing qualifier
    ("אשכול    (מ-1 עד 10)"), so prefix matching keeps ingestion robust to small
    formatting drift in the source workbook.
    """
    mapping: dict[str, str] = {}
    for raw in raw_columns:
        key = _norm_header(raw)
        for heb, canon in rename_cfg.items():
            if key == heb or key.startswith(heb):
                mapping[raw] = canon
                break
    return mapping


def load_ses(cfg: dict[str, Any] | None = None) -> pd.DataFrame:
    """Load the CBS socioeconomic index workbook."""
    cfg = cfg or load_config()
    s = cfg["ses"]
    df = pd.read_excel(resolve(cfg["paths"]["raw_ses"]),
                       sheet_name=s["sheet"], header=s["header_row"])
    df = df.dropna(how="all").reset_index(drop=True)

    df = df.rename(columns=_resolve_rename_prefix(list(df.columns), s["rename"]))

    # CBS "unranked" placeholders -> NaN, then coerce the numerics.
    df = df.replace({tok: pd.NA for tok in s["na_tokens"]})
    for col in s["numeric_columns"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    if "locality_name" in df.columns:
        df = df[df["locality_name"].notna()].reset_index(drop=True)
    return df


# --------------------------------------------------------------------------- #
# Dataset 3 — Ministry of Education budget                                      #
# --------------------------------------------------------------------------- #
def _patch_openpyxl_colors() -> None:
    """Make openpyxl tolerate the budget workbook's invalid theme colours.

    Without this, ``load_workbook`` aborts with
    ``ValueError: Colors must be aRGB hex values`` before a single cell is read.
    """
    import openpyxl.styles.colors as colors

    if getattr(colors.RGB, "_v2_patched", False):
        return
    _orig = colors.RGB.__set__

    def _lenient(self, instance, value):  # noqa: ANN001
        try:
            _orig(self, instance, value)
        except (ValueError, TypeError):
            _orig(self, instance, "00000000")

    colors.RGB.__set__ = _lenient
    colors.RGB._v2_patched = True


def load_budget(cfg: dict[str, Any] | None = None) -> tuple[pd.DataFrame, dict[str, Any]]:
    """Load the budget workbook; return (raw selected frame, ingestion report).

    Steps:
      1. Patch openpyxl, open ``read_only`` + ``data_only`` (values, not formulae).
      2. Whitespace-normalise the Hebrew headers.
      3. Drop the grand-totals row (any cell equal to ``סה"כ``).
      4. Rename the configured columns EXACTLY (prefix matching would confuse
         'סה"כ תקציב שכר ותשלומים' with its '- ללא קורונה' sibling) and keep only
         those we actually use.
    """
    cfg = cfg or load_config()
    b = cfg["budget"]
    _patch_openpyxl_colors()
    import openpyxl

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(resolve(cfg["paths"]["raw_budget"]),
                                    read_only=True, data_only=True)
        ws = wb[b["sheet"]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    header = [_norm_header(c) for c in rows[b["header_row"]]]
    df = pd.DataFrame(rows[b["header_row"] + 1:], columns=header)
    n_raw = len(df)

    # Drop the grand-totals row.
    totals = b["totals_label"]
    is_totals = df.apply(lambda r: r.astype(str).str.strip().eq(totals).any(), axis=1)
    df = df[~is_totals].reset_index(drop=True)

    # EXACT rename of the configured subset.
    rename_cfg: dict[str, str] = b["rename"]
    present = {h: canon for h, canon in rename_cfg.items() if h in df.columns}
    missing = [h for h in rename_cfg if h not in df.columns]
    df = df[list(present.keys())].rename(columns=present)

    report = {
        "rows_raw": n_raw,
        "rows_after_totals_drop": len(df),
        "totals_rows_dropped": int(is_totals.sum()),
        "columns_in_workbook": len(header),
        "columns_requested": len(rename_cfg),
        "columns_resolved": len(present),
        "columns_missing": missing,
        "known_empty_excluded": b.get("known_empty_columns", []),
    }
    return df, report


if __name__ == "__main__":
    cfg = load_config()
    bag = load_bagrut(cfg)
    ses = load_ses(cfg)
    bud, rep = load_budget(cfg)
    print(f"[io_load] bagrut: {bag.shape}")
    print(f"[io_load] ses:    {ses.shape}")
    print(f"[io_load] budget: {bud.shape}  report={rep}")
